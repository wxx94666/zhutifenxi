# -*- coding: utf-8 -*-
import pandas as pd
import numpy as np
import streamlit as st
import plotly.express as px
import plotly.graph_objects as go
from openpyxl import load_workbook
import plotly.io as pio
import os

# 设置Plotly默认主题
pio.templates.default = "plotly_white"

# ========= 新增：是否在各自页面内显示“数据更新时间”的开关（已统一挪到顶部，故默认不展示） =========
SHOW_TS_INLINE = False

# ========= 基础工具 =========
LETTER_IDX = {c: i for i, c in enumerate(list("ABCDEFGHIJKLMNOPQRSTUVWXYZ"))}

def col(df, letter):
    return df.iloc[:, LETTER_IDX[letter]]

def safe_sum(series):
    return float(pd.to_numeric(series, errors="coerce").fillna(0).sum())

def to_pct(x, denom):
    try:
        return (float(x) / float(denom)) * 100 if denom else 0.0
    except Exception:
        return 0.0

def coerce_numeric(series, treat_percent=False):
    s = series.astype(str).str.strip().replace({"None": np.nan, "nan": np.nan})
    s = s.str.replace("%", "", regex=False)
    vals = pd.to_numeric(s, errors="coerce")
    if treat_percent and (vals.dropna() > 1).any():
        vals = vals / 100.0
    return vals.fillna(0)

def group_metric(df, group_letter, value_letter, how="mean"):
    g = df.groupby(col(df, group_letter))
    if how == "sum":
        s = g.apply(lambda x: pd.to_numeric(col(x, value_letter), errors="coerce").fillna(0).sum())
    elif how == "max":
        s = g.apply(lambda x: pd.to_numeric(col(x, value_letter), errors="coerce").fillna(0).max())
    else:
        s = g.apply(lambda x: pd.to_numeric(col(x, value_letter), errors="coerce").dropna().mean())
    s = s.replace([np.inf, -np.inf], np.nan).fillna(0)
    return s.reset_index().rename(columns={0: "value", s.index.name: "主体"})

# ========= 读取 X1 =========
def read_x1_text(excel_path):
    candidates = ["数据表", "主体分析表"]
    try:
        wb = load_workbook(excel_path, data_only=True)
        for name in candidates:
            if name in wb.sheetnames:
                v = wb[name]["X1"].value
                if v is not None:
                    return str(v).strip()
    except Exception:
        pass
    for name in candidates:
        try:
            df_raw = pd.read_excel(excel_path, sheet_name=name, header=None)
            return str(df_raw.iat[0, LETTER_IDX["X"]]).strip()
        except Exception:
            continue
    return "未提供"

# ========= KPI（主体页）=========
def build_kpis_subject(df_subject, excel_path):
    ts_str = read_x1_text(excel_path)
    total_accounts = safe_sum(col(df_subject, "C"))
    delivered_accounts = safe_sum(col(df_subject, "D"))
    unbound_cards = safe_sum(col(df_subject, "F"))
    dead_count = safe_sum(col(df_subject, "H"))
    avg_dead_rate = to_pct(dead_count, total_accounts)
    return ts_str, total_accounts, delivered_accounts, unbound_cards, avg_dead_rate

# ========= TopN 横向柱状图（主体页通用）=========
def topn_chart(df_subject, group_letter, value_letter, title, how="mean", n=10, is_percent=False, sort_ascending=False):
    tmp = df_subject.copy()
    tmp.iloc[:, LETTER_IDX[value_letter]] = coerce_numeric(col(tmp, value_letter), treat_percent=is_percent)
    agg = group_metric(tmp, group_letter, value_letter, how=how)
    agg = agg.rename(columns={agg.columns[0]: "主体", agg.columns[1]: "值"}).copy()
    agg["值"] = pd.to_numeric(agg["值"], errors="coerce").fillna(0)

    agg = agg.sort_values("值", ascending=sort_ascending).head(n)
    order = agg["主体"].tolist()

    if is_percent:
        text_vals = (agg["值"] * 100).round(2).astype(str) + "%"
        x_vals = agg["值"]
    else:
        text_vals = agg["值"].round(2).astype(str)
        x_vals = agg["值"]

    fig = px.bar(
        agg, x=x_vals, y="主体", orientation="h",
        title=title, text=text_vals,
        color="值",
        color_continuous_scale="Viridis",
        height=450
    )
    fig.update_traces(
        textposition="inside",
        textangle=0,
        insidetextanchor="middle",
        marker_line_width=0
    )
    fig.update_layout(
        yaxis=dict(categoryorder="array", categoryarray=order, autorange="reversed"),
        margin=dict(l=150, r=30, t=70, b=40),
        coloraxis_showscale=False,
        title_font=dict(size=18),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        hovermode="y unified"
    )
    return fig

# ========= 饼图（主体页）=========
def pie_from_filter(df_subject, filter_value, group_letter, value_letter, title, percent_mode=False, total_denominator=None):
    sub = df_subject[col(df_subject, "A") == filter_value].copy()
    sub.iloc[:, LETTER_IDX[value_letter]] = coerce_numeric(col(sub, value_letter), treat_percent=False)
    grouped = sub.groupby(col(sub, group_letter)).apply(lambda x: x.iloc[:, LETTER_IDX[value_letter]].sum())
    grouped = grouped.replace([np.inf, -np.inf], np.nan).fillna(0)
    dfp = grouped.reset_index()
    dfp.columns = ["主体", "值"]

    dfp = dfp[dfp["值"] > 0]

    if percent_mode:
        denom = float(total_denominator) if total_denominator else 0.0
        dfp["值"] = dfp["值"].astype(float).apply(lambda v: (v / denom * 100.0) if denom > 0 else 0.0)

    fig = px.pie(
        dfp,
        values="值",
        names="主体",
        title=title,
        hole=0.4,
        color_discrete_sequence=px.colors.qualitative.Pastel
    )
    fig.update_traces(
        textinfo="label+percent",
        textposition="inside",
        insidetextorientation="radial",
        marker=dict(line=dict(color="#FFFFFF", width=2))
    )
    fig.update_layout(
        title_font=dict(size=18),
        margin=dict(l=20, r=20, t=70, b=20),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        legend=dict(
            orientation="h",
            yanchor="bottom",
            y=1.02,
            xanchor="right",
            x=1
        )
    )
    return fig

# ========= 账户排行榜（主体页）=========
def account_rank_table(df_account):
    status = col(df_account, "H").astype(str).str.strip().str.lower()
    mask = status.isin(["active", "need to pay"])
    sub = df_account[mask].copy()
    sub["排序值"] = pd.to_numeric(col(sub, "J"), errors="coerce")
    sub = sub.sort_values("排序值", ascending=False).head(10)
    out = pd.DataFrame({
        "排名": range(1, len(sub)+1),
        "客户名称": col(sub, "C").astype(str),
        "账户名称": col(sub, "D").astype(str),
        "账户ID": col(sub, "E").astype(str),
        "一周消耗": col(sub, "J")
    })
    return out.set_index("排名")

# ========= 所有主体状态详情 =========
def all_subjects_status(df_subject):
    sub = df_subject.copy()
    for l in ["B","C","D","E","F","G","H"]:
        if l != "B":
            sub[l] = coerce_numeric(col(sub, l), treat_percent=False)
        else:
            sub[l] = col(sub, "B").astype(str)

    agg = sub.groupby("B").agg({
        "C":"sum","D":"sum","E":"sum","F":"sum","G":"sum","H":"sum"
    }).reset_index().rename(columns={"B":"主体"})

    totals = agg["C"].astype(float)
    total_states = agg["D"]+agg["E"]+agg["F"]+agg["G"]+agg["H"]
    overflow = (total_states - totals).clip(lower=0)
    agg["G"] = (agg["G"] - overflow).clip(lower=0)

    denom = totals.replace(0, np.nan)
    pct_cols = {}
    for l in ["D","E","F","G","H"]:
        pct_cols[l] = (agg[l] / denom * 100).fillna(0)

    stacked_sum = pct_cols["D"] + pct_cols["E"] + pct_cols["F"] + pct_cols["G"] + pct_cols["H"]
    filler_pct = (100 - stacked_sum).clip(lower=0)

    cats = ["D","E","F","G","H"]
    names = {"D":"已交付","E":"可交付","F":"未绑卡","G":"问题户","H":"死户"}
    colors = {"D":"#3498db","E":"#2ecc71","F":"#f39c12","G":"#9b59b6","H":"#e74c3c"}

    agg = agg.assign(
        D_pct=pct_cols["D"], E_pct=pct_cols["E"], F_pct=pct_cols["F"],
        G_pct=pct_cols["G"], H_pct=pct_cols["H"], FILLER=filler_pct
    )
    agg = agg.sort_values("C", ascending=False).reset_index(drop=True)
    order = agg["主体"].tolist()

    fig = go.Figure()
    for c in cats:
        fig.add_trace(go.Bar(
            y=agg["主体"],
            x=agg[f"{c}_pct"],
            orientation="h",
            name=names[c],
            marker=dict(color=colors[c]),
            text=agg[c].astype(int),
            textposition="inside",
            insidetextanchor="middle",
            textangle=0,
            cliponaxis=False,
            marker_line_width=0
        ))

    fig.add_trace(go.Bar(
        y=agg["主体"],
        x=agg["FILLER"],
        orientation="h",
        name="",
        marker=dict(color="rgba(0,0,0,0)"),
        hoverinfo="skip",
        showlegend=False
    ))

    fig.add_trace(go.Scatter(
        y=agg["主体"], x=[110]*len(agg), mode="text",
        text=[f"<b>{int(v)}</b>" for v in agg["C"]],
        textposition="middle left",
        textfont=dict(color="#333333", size=14),
        showlegend=False, hoverinfo="skip", cliponaxis=False
    ))

    fig.add_annotation(
        x=110, y=1.02, xref="x", yref="paper",
        text="<b>总账户数</b>", showarrow=False,
        font=dict(color="#333333", size=14),
        xanchor="left", yanchor="bottom"
    )

    fig.update_layout(
        barmode="stack",
        title="所有主体状态详情",
        title_font=dict(size=18),
        xaxis=dict(showticklabels=False, showgrid=False, zeroline=False, range=[0,125]),
        yaxis=dict(title="主体", categoryorder="array", categoryarray=order),
        margin=dict(l=150, r=180, t=70, b=40),
        height=max(550, 30*len(agg)),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        legend=dict(orientation="h", yanchor="bottom", y=1.04, xanchor="center", x=0.5),
        hovermode="y unified"
    )
    return fig

# ========= 客户页：工具 =========
def build_kpis_customer(df_customer):
    total_clients = df_customer.shape[0]
    total_accounts = safe_sum(col(df_customer, "B"))
    delivered_total = safe_sum(col(df_customer, "E"))
    dead_total = safe_sum(col(df_customer, "G"))
    total_consumption = safe_sum(col(df_customer, "I"))
    return total_clients, total_accounts, delivered_total, dead_total, total_consumption

def customer_topn_bar(df_customer, value_letter, title, n=10, is_percent=False, ascending=False):
    df = df_customer.copy()
    names = col(df, "A").astype(str)
    vals = coerce_numeric(col(df, value_letter), treat_percent=is_percent)
    out = pd.DataFrame({"客户": names, "值": vals})
    out = out.sort_values("值", ascending=ascending).head(n)

    text_vals = (out["值"]*100).round(2).astype(str) + "%" if is_percent else out["值"].round(2).astype(str)

    fig = px.bar(
        out, x="值", y="客户", orientation="h",
        title=title, text=text_vals,
        color="值", color_continuous_scale="Viridis", height=450
    )
    fig.update_traces(textposition="inside", insidetextanchor="middle", textangle=0, marker_line_width=0)
    fig.update_layout(
        yaxis=dict(autorange="reversed"),
        margin=dict(l=150, r=30, t=70, b=40),
        coloraxis_showscale=False,
        title_font=dict(size=18),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        hovermode="y unified"
    )
    return fig

def customer_pie_internal_external(df_customer):
    internal_sum = safe_sum(col(df_customer, "C"))
    external_sum = safe_sum(col(df_customer, "D"))
    dfp = pd.DataFrame({"类型":["内部户","外部户"], "值":[internal_sum, external_sum]})
    fig = px.pie(dfp, values="值", names="类型", title="客户内部/外部户使用占比", hole=0.4,
                 color_discrete_sequence=px.colors.qualitative.Pastel)
    fig.update_traces(textinfo="label+percent", textposition="inside",
                      marker=dict(line=dict(color="#fff", width=2)))
    fig.update_layout(margin=dict(l=20,r=20,t=70,b=20))
    return fig

def customer_pie_account_types(df_customer):
    e = safe_sum(col(df_customer, "E"))
    f = safe_sum(col(df_customer, "F"))
    g = safe_sum(col(df_customer, "G"))
    dfp = pd.DataFrame({"类别":["已交付","问题户","死户"], "值":[e,f,g]})
    fig = px.pie(dfp, values="值", names="类别", title="客户账户类型占比", hole=0.4,
                 color_discrete_sequence=px.colors.qualitative.Pastel)
    fig.update_traces(textinfo="label+percent", textposition="inside",
                      marker=dict(line=dict(color="#fff", width=2)))
    fig.update_layout(margin=dict(l=20,r=20,t=70,b=20))
    return fig

def customer_pie_consumption_ratio(df_customer):
    vals = coerce_numeric(col(df_customer, "I"), treat_percent=False)
    gt3 = int((vals > 3).sum())
    le3 = int((vals <= 3).sum())
    dfp = pd.DataFrame({"分组":["> 3","≤ 3"], "数量":[gt3, le3]})
    fig = px.pie(dfp, values="数量", names="分组", title="客户消耗占比（按阈值3计数）", hole=0.4,
                 color_discrete_sequence=px.colors.qualitative.Pastel)
    fig.update_traces(textinfo="label+percent", textposition="inside",
                      marker=dict(line=dict(color="#fff", width=2)))
    fig.update_layout(margin=dict(l=20,r=20,t=70,b=20))
    return fig

def customer_pie_accounts_share(df_customer):
    names_raw = col(df_customer, "A").astype(str)
    mask_valid_name = names_raw.fillna("").str.strip().replace({"nan":""}).astype(str)
    mask_valid = mask_valid_name != ""
    names = names_raw[mask_valid]
    totals = coerce_numeric(col(df_customer, "B"))[mask_valid]

    dfp = pd.DataFrame({"客户": names.values, "账户总数": totals.values})
    dfp = dfp[dfp["账户总数"] > 0]
    fig = px.pie(dfp, values="账户总数", names="客户", title="客户账户总数占比", hole=0.4,
                 color_discrete_sequence=px.colors.qualitative.Pastel)
    fig.update_traces(textinfo="label+percent", textposition="inside",
                      marker=dict(line=dict(color="#fff", width=2)))
    fig.update_layout(margin=dict(l=20,r=20,t=70,b=20))
    return fig

def customer_consumption_rank_table(df_customer):
    name = col(df_customer, "A").astype(str)
    delivered = coerce_numeric(col(df_customer, "E"))
    dead_rate = coerce_numeric(col(df_customer, "H"), treat_percent=True)   # 小数(0~1)
    total_cons = coerce_numeric(col(df_customer, "I"))
    avg_cons = coerce_numeric(col(df_customer, "J"))

    df = pd.DataFrame({
        "客户名称": name,
        "已交付数量": delivered,
        "客户死户率": (dead_rate * 100).round(2),
        "客户总消耗": total_cons.round(2),
        "客户平均消耗": avg_cons.round(2)
    })
    df = df.sort_values("客户总消耗", ascending=False).head(10).reset_index(drop=True)
    df.insert(0, "序号", range(1, len(df) + 1))
    df = df.loc[:, ~df.columns.duplicated()].copy()
    df = df[["序号", "客户名称", "已交付数量", "客户死户率", "客户总消耗", "客户平均消耗"]]
    return df

def all_customers_status(df_customer):
    df = df_customer.copy()
    names = col(df, "A").astype(str)
    B = coerce_numeric(col(df, "B"))
    E = coerce_numeric(col(df, "E"))
    F = coerce_numeric(col(df, "F"))
    G = coerce_numeric(col(df, "G"))

    denom = B.replace(0, np.nan)
    E_pct = (E / denom * 100).fillna(0)
    F_pct = (F / denom * 100).fillna(0)
    G_pct = (G / denom * 100).fillna(0)
    stacked = E_pct + F_pct + G_pct
    filler = (100 - stacked).clip(lower=0)

    fig = go.Figure()
    colors = {"E":"#3498db", "F":"#9b59b6", "G":"#e74c3c"}
    labels = {"E":"已交付", "F":"问题户", "G":"死户"}

    fig.add_trace(go.Bar(y=names, x=E_pct, name=labels["E"], orientation="h",
                         marker=dict(color=colors["E"]), text=E.astype(int),
                         textposition="inside", insidetextanchor="middle", marker_line_width=0))
    fig.add_trace(go.Bar(y=names, x=F_pct, name=labels["F"], orientation="h",
                         marker=dict(color=colors["F"]), text=F.astype(int),
                         textposition="inside", insidetextanchor="middle", marker_line_width=0))
    fig.add_trace(go.Bar(y=names, x=G_pct, name=labels["G"], orientation="h",
                         marker=dict(color=colors["G"]), text=G.astype(int),
                         textposition="inside", insidetextanchor="middle", marker_line_width=0))
    fig.add_trace(go.Bar(y=names, x=filler, name="", orientation="h",
                         marker=dict(color="rgba(0,0,0,0)"), hoverinfo="skip", showlegend=False))

    fig.add_trace(go.Scatter(
        y=names, x=[110]*len(names), mode="text",
        text=[f"<b>{int(v)}</b>" for v in B],
        textposition="middle left", textfont=dict(color="#333333", size=14),
        showlegend=False, hoverinfo="skip", cliponaxis=False
    ))
    fig.add_annotation(
        x=110, y=1.02, xref="x", yref="paper",
        text="<b>总账户数</b>", showarrow=False,
        font=dict(color="#333333", size=14),
        xanchor="left", yanchor="bottom"
    )

    fig.update_layout(
        barmode="stack",
        title="所有客户状态详情",
        title_font=dict(size=18),
        xaxis=dict(showticklabels=False, showgrid=False, zeroline=False, range=[0,125]),
        yaxis=dict(title="客户", autorange="reversed"),
        margin=dict(l=150, r=180, t=70, b=40),
        height=max(550, 28*len(names)),
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        legend=dict(orientation="h", yanchor="bottom", y=1.04, xanchor="center", x=0.5),
        hovermode="y unified"
    )
    return fig

# ========= 页面渲染：主体 =========
def render_subject_page(df_subject, df_account, excel_path):
    st.markdown('<div class="big-title">Lucky主体分析看板</div>', unsafe_allow_html=True)
    ts_str, total_accounts, delivered_accounts, unbound_cards, avg_dead_rate = build_kpis_subject(df_subject, excel_path)
    # 统一显示在顶部，此处不再显示
    if SHOW_TS_INLINE:
        st.markdown(f'<div class="data-time">数据更新时间：{ts_str}</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-header">数据概览</div>', unsafe_allow_html=True)
    with st.container():
        cols = st.columns(5)
        metrics = [
            ("总账户数", f"{int(total_accounts):,}"),
            ("已交付账户数", f"{int(delivered_accounts):,}"),
            ("未绑卡账户数", f"{int(unbound_cards):,}"),
            ("平均死户率", f"{avg_dead_rate:.2f}%"),
            ("账户活跃度", f"{to_pct(delivered_accounts, total_accounts):.2f}%")
        ]
        for i, (label, value) in enumerate(metrics):
            with cols[i]:
                st.markdown('<div class="metric-container">', unsafe_allow_html=True)
                st.metric(label, value)
                st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-header">榜单 Top10</div>', unsafe_allow_html=True)
    c1, c2 = st.columns(2)
    with c1:
        st.plotly_chart(
            topn_chart(df_subject, "B", "R", "主体综合评分榜单（R）", how="max", n=10, is_percent=False, sort_ascending=False),
            use_container_width=True
        )
        st.plotly_chart(
            topn_chart(df_subject, "B", "L", "主体平均消耗榜单（L）", how="mean", n=10, is_percent=False),
            use_container_width=True
        )
    with c2:
        st.plotly_chart(
            topn_chart(df_subject, "B", "P", "主体好户榜单（P）", how="max", n=10, is_percent=True),
            use_container_width=True
        )
        st.plotly_chart(
            topn_chart(df_subject, "B", "I", "主体死户率榜单（I）- 最低10名", how="max", n=10, is_percent=True, sort_ascending=True),
            use_container_width=True
        )

    st.markdown('<div class="section-header">占比分析</div>', unsafe_allow_html=True)
    total_accounts_all = safe_sum(col(df_subject, "C"))
    c3, c4 = st.columns(2)
    with c3:
        st.plotly_chart(pie_from_filter(df_subject, "内部户表", "B", "C", "内部户主体账户数占比"), use_container_width=True)
        st.plotly_chart(pie_from_filter(df_subject, "内部户表", "B", "O", "内部户主体消耗占比"), use_container_width=True)
        st.plotly_chart(
            pie_from_filter(df_subject, "内部户表", "B", "H", "内部户主体死户数占比（占总账户数%）",
                            percent_mode=True, total_denominator=total_accounts_all),
            use_container_width=True
        )
    with c4:
        st.plotly_chart(pie_from_filter(df_subject, "外部户表", "B", "C", "外部户主体账户数占比"), use_container_width=True)
        st.plotly_chart(pie_from_filter(df_subject, "外部户表", "B", "O", "外部户主体消耗占比"), use_container_width=True)
        st.plotly_chart(
            pie_from_filter(df_subject, "外部户表", "B", "H", "外部户主体死户数占比（占总账户数%）",
            percent_mode=True, total_denominator=total_accounts_all),
            use_container_width=True
        )

    st.markdown('<div class="section-header">账户排行榜（Top 10）</div>', unsafe_allow_html=True)
    rank_df = account_rank_table(df_account)
    st.dataframe(
        rank_df,
        use_container_width=True,
        column_config={"一周消耗": st.column_config.NumberColumn(format="%.2f")}
    )

    st.markdown('<div class="section-header">所有主体状态详情</div>', unsafe_allow_html=True)
    st.plotly_chart(all_subjects_status(df_subject), use_container_width=True)

# ========= 页面渲染：客户 =========
def render_customer_page(df_customer, excel_path):
    st.markdown('<div class="big-title">Lucky客户分析看板</div>', unsafe_allow_html=True)
    ts_str = read_x1_text(excel_path)
    # 统一显示在顶部，此处不再显示
    if SHOW_TS_INLINE:
        st.markdown(f'<div class="data-time">数据更新时间：{ts_str}</div>', unsafe_allow_html=True)

    total_clients, total_accounts, delivered_total, dead_total, total_consumption = build_kpis_customer(df_customer)
    st.markdown('<div class="section-header">数据概览</div>', unsafe_allow_html=True)
    with st.container():
        cols = st.columns(5)
        metrics = [
            ("客户总数", f"{int(total_clients):,}"),
            ("客户账户总数", f"{int(total_accounts):,}"),
            ("客户交付总数", f"{int(delivered_total):,}"),
            ("客户死户总数", f"{int(dead_total):,}"),
            ("客户总消耗", f"{total_consumption:.2f}")
        ]
        for i, (label, value) in enumerate(metrics):
            with cols[i]:
                st.markdown('<div class="metric-container">', unsafe_allow_html=True)
                st.metric(label, value)
                st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('<div class="section-header">榜单 Top10</div>', unsafe_allow_html=True)
    a1, a2 = st.columns(2)
    with a1:
        st.plotly_chart(
            customer_topn_bar(df_customer, "O", "客户综合评分榜单（O）", n=10, is_percent=False, ascending=False),
            use_container_width=True
        )
        st.plotly_chart(
            customer_topn_bar(df_customer, "E", "客户交付数量榜（E）", n=10, is_percent=False, ascending=False),
            use_container_width=True
        )
    with a2:
        st.plotly_chart(
            customer_topn_bar(df_customer, "I", "客户消耗榜（I）", n=10, is_percent=False, ascending=False),
            use_container_width=True
        )
        tmp = df_customer.copy()
        hvals = coerce_numeric(col(tmp, "H"), treat_percent=True)  # 0~1
        tmp = tmp[hvals < 1]
        st.plotly_chart(
            customer_topn_bar(tmp, "H", "客户死户率榜（H）- 最高10名", n=10, is_percent=True, ascending=False),
            use_container_width=True
        )

    st.markdown('<div class="section-header">占比分析</div>', unsafe_allow_html=True)
    b1, b2 = st.columns(2)
    with b1:
        st.plotly_chart(customer_pie_internal_external(df_customer), use_container_width=True)
        st.plotly_chart(customer_pie_account_types(df_customer), use_container_width=True)
    with b2:
        st.plotly_chart(customer_pie_consumption_ratio(df_customer), use_container_width=True)
        st.plotly_chart(customer_pie_accounts_share(df_customer), use_container_width=True)

    st.markdown('<div class="section-header">客户消耗排行榜（Top 10）</div>', unsafe_allow_html=True)
    table_df = customer_consumption_rank_table(df_customer)
    st.dataframe(
        table_df,
        use_container_width=True,
        hide_index=True,
        column_config={
            "序号": st.column_config.NumberColumn(format="%d"),
            "客户死户率": st.column_config.NumberColumn(format="%.2f%%"),
            "客户总消耗": st.column_config.NumberColumn(format="%.2f"),
            "客户平均消耗": st.column_config.NumberColumn(format="%.2f")
        }
    )

    st.markdown('<div class="section-header">所有客户状态详情</div>', unsafe_allow_html=True)
    st.plotly_chart(all_customers_status(df_customer), use_container_width=True)

# ========= 页面渲染：主体详情表（第3页）=========
def render_subject_detail_page(df_subject):
    st.markdown('<div class="big-title">主体详情表</div>', unsafe_allow_html=True)
    st.dataframe(df_subject, use_container_width=True)

# ========= 页面渲染：客户详情表（第4页）=========
def render_customer_detail_page(df_customer):
    st.markdown('<div class="big-title">客户详情表</div>', unsafe_allow_html=True)
    df_disp = df_customer.copy()
    for cname in ["客户死户率", "账户消耗占比"]:
        if cname in df_disp.columns:
            df_disp[cname] = coerce_numeric(df_disp[cname], treat_percent=True) * 100
    st.dataframe(
        df_disp,
        use_container_width=True,
        column_config={
            "客户死户率": st.column_config.NumberColumn(format="%.2f%%"),
            "账户消耗占比": st.column_config.NumberColumn(format="%.2f%%"),
        }
    )

# ========= 页面渲染：客户主体文本数据（第5页）=========
def render_text_page():
    import os
    FILE_NAME = "主体与客户分析文本数据.txt"

    st.markdown('<div class="big-title">客户主体文本数据</div>', unsafe_allow_html=True)

    if not os.path.exists(FILE_NAME):
        st.warning("文本文件未上传。")
        return

    with open(FILE_NAME, "r", encoding="utf-8") as f:
        content = f.read()

    # 按标记分割
    part1, part2, part3 = "", "", ""
    if "【主体账户剩余统计】" in content:
        idx1 = content.find("【主体账户剩余统计】")
        idx2 = content.find("【主体分析表数据】")
        idx3 = content.find("【客户分析表数据】")
        if idx2 > 0:
            part1 = content[idx1:idx2].strip()
        if idx3 > 0:
            part2 = content[idx2:idx3].strip()
            part3 = content[idx3:].strip()

    # 三列布局
    c1, c2, c3 = st.columns(3)
    with c1:
        st.markdown(
            f"<div class='text-box'><pre>{part1}</pre></div>",
            unsafe_allow_html=True
        )
    with c2:
        st.markdown(
            f"<div class='text-box'><pre>{part2}</pre></div>",
            unsafe_allow_html=True
        )
    with c3:
        st.markdown(
            f"<div class='text-box'><pre>{part3}</pre></div>",
            unsafe_allow_html=True
        )

# ========= Streamlit App =========
st.set_page_config(page_title="分析可视化", layout="wide")

# ========= 新增：全局样式（含右上角时间条样式）=========
st.markdown(
    """
    <style>
   /* ---- 文本三列页面样式 ---- */
.txt-col {
  background: #ffffff;
  border: 1px solid #e6e8eb;
  border-radius: 10px;
  padding: 14px 16px;
  box-shadow: 0 2px 5px rgba(0,0,0,0.05);
}
.txt-col .txt-title {
  font-weight: 600;
  color: #111111;
  margin-bottom: 10px;
}
.txt-col pre {
  white-space: pre;
  word-wrap: normal;
  overflow-x: auto;
  overflow-y: auto;
  margin: 0;
  font-family: ui-monospace, SFMono-Regular, Menlo, Monaco, Consolas, "Liberation Mono", "Courier New", monospace;
  font-size: 14px;
  line-height: 1.6;
  color: #111111;
  background: #ffffff;
  border-radius: 6px;
  padding: 10px 12px;
  border: 1px solid #eef0f2;
}
.text-box {
    background-color: #000000;
    color: #ffffff;
    padding: 15px;
    border-radius: 8px;
    white-space: pre-wrap;
    font-family: Consolas, monospace;
    font-size: 14px;
    line-height: 1.5;
}
.top-right-ts {
    text-align: right;
    margin: 4px 2px 8px 2px;
    font-size: 14px;
    color: #7f8c8d;
}
    </style>
    """, unsafe_allow_html=True
)

# ========= 新增：简单密码访问（会话级）=========
if "authed" not in st.session_state:
    st.session_state.authed = False

if not st.session_state.authed:
    st.markdown("### 请输入访问密码")
    with st.form("auth_form", clear_on_submit=True):
        # 去掉占位提示词（placeholder）
        pwd = st.text_input("密码", type="password")
        ok = st.form_submit_button("进入")
        if ok:
            if pwd == "lucky":
                st.session_state.authed = True
                # 使用新版 API，兼容旧版做后备
                try:
                    st.rerun()
                except AttributeError:
                    st.experimental_rerun()
            else:
                st.error("密码错误，请重试。")
    st.stop()


# 读取数据
EXCEL_FILE = "主体分析.xlsx"
try:
    df_subject = pd.read_excel(EXCEL_FILE, sheet_name="主体分析表", header=0)
    df_account = pd.read_excel(EXCEL_FILE, sheet_name="账户表", header=0)
except Exception as e:
    st.error(f"主体数据加载失败: {str(e)}")
    st.stop()

try:
    df_customer = pd.read_excel(EXCEL_FILE, sheet_name="客户分析表", header=0)
except Exception as e:
    st.error(f"客户数据加载失败: {str(e)}")
    st.stop()

# ========= 新增：顶部统一显示“数据更新时间”（靠右）=========
_ts_global = read_x1_text(EXCEL_FILE)
st.markdown(f"<div class='top-right-ts'>数据更新时间：{_ts_global}</div>", unsafe_allow_html=True)

# 顶部标签页（5个）
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "Lucky主体分析看板",
    "Lucky客户分析看板",
    "主体详情表",
    "客户详情表",
    "客户主体文本数据"
])

with tab1:
    render_subject_page(df_subject, df_account, EXCEL_FILE)
with tab2:
    render_customer_page(df_customer, EXCEL_FILE)
with tab3:
    render_subject_detail_page(df_subject)
with tab4:
    render_customer_detail_page(df_customer)
with tab5:
    render_text_page()

# 页脚
st.markdown(
    """
    <hr>
    <div style="text-align: center; color: #7f8c8d; font-size: 14px;">
        Lucky主体/客户分析可视化系统 © 2025
    </div>
    """,
    unsafe_allow_html=True
)
